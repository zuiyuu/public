from asyncio.windows_events import NULL
from openpyxl import load_workbook
import  os
from itertools import zip_longest

rowNum = 0
colNum = 0
backUpDirectory = ''
excel = NULL

def setFilePath (filePath) :
    if (filePath != '') :
        return filePath

def getSheetNames(excel) :
    # book = xlrd.open_workbook(backUpDirectory)
    print("The number of worksheets is {0}".format(len(excel.sheetnames)))
    print("Worksheet name(s): {0}".format(excel.sheetnames))
    return excel.sheetnames

def getSheetDataList(excel,sheetName) :
    sqlList = []
    table = excel.get_sheet_by_name(sheetName)
    # Data=table.cell(row=rowNum,column=colNum).value
    for i in range(3,table.max_row+1): #从2开始，第一行为title
        #遍历Excel表格列的内容
        sqlData = (table.cell(row=i,column=1)).value #获取数据
        if ((sqlData != '') and not(sqlData is None)) :
            sqlList.append(sqlData)
            print("读取到的SQL文：%s",sqlData)
    return sqlList

# 指定规则：表项目取得
# 返回值  1:读取Excel Sheet的一个位置的表明 String
#        2：读取Excel Sheet的第一行的各项目名 List
def getSheetHeadDataList(excel,sheetName) :
    sqlTableName = ''
    sqlHead = []
    table = excel.get_sheet_by_name(sheetName)
    sqlTableName = (table.cell(row=1,column=1)).value #获取数据
    # Data=table.cell(row=rowNum,column=colNum).value
    for i in range(4,table.max_column+1): #从3开始，第一行为title
        #遍历Excel表格列的内容
        sqlData = (table.cell(row=2,column=i)).value #获取数据
        if ((sqlData != '') and not(sqlData is None)) :
            sqlHead.append(sqlData)
            # print("读取到的SQL项目名：%s",sqlHead)
    return sqlTableName,sqlHead

# 指定规则：各项目确认内容取得
def getSheetBodyDataList(excel,sheetName) :
    sqlBodyList = []
    table = excel.get_sheet_by_name(sheetName)
    # nullCheckList = [False,False]
    for x in range(3,table.max_row+1): #从2开始，第一行为title
    
        sqlBodyAllList = []
        # Data=table.cell(row=rowNum,column=colNum).value
        nullCheck = False
        for i in range(4,table.max_column+1): #从3开始，第一行为title
            #遍历Excel表格列的内容
            sqlData = (table.cell(row=x,column=i)).value #获取数据
            if (not(sqlData is None)) :
                nullCheck = True
                sqlBodyAllList.append(sqlData)
                # print("读取到的SQL文各项目值：%s",sqlBodyAllList)
            else:
                # Excel空的时候，值为None
                sqlBodyAllList.append('')
                # nullCheckList.append()
                # print("读取到的SQL文各项目值错误")
        # 一行的数据不全为空的数据
        if (nullCheck) :
            sqlBodyList.append(sqlBodyAllList)
    return sqlBodyList
    
def is_kazu(num):
    if type(num) is not str:
        raise ValueError('parameter must be a string.')
    if len(num) > 2 and num.count('.',1,-1) == 1:
        num=num.replace('.','',1)
    if num.isnumeric():
        return True
    else:
        return False
    
def openExecl(paraUpDirectory):
    return load_workbook(paraUpDirectory, data_only=True)

def getInputExcel(excel, sheetName):
    sqlBodyList = []
    table = excel.get_sheet_by_name(sheetName)
    # nullCheckList = [False,False]
    for x in range(1,table.max_row+1): #从0开始，第一行为title
    
        sqlBodyAllList = {}
        # Data=table.cell(row=rowNum,column=colNum).value
        nullCheck = False
        for i in range(1,table.max_column+1): #从0开始，第一行为title
            #遍历Excel表格列的内容
            ruleData = (table.cell(row=x,column=i)).value #获取数据
            if (not(ruleData is None)) :
                nullCheck = True
                sqlBodyAllList.append(ruleData)
                # print("读取到的SQL文各项目值：%s",sqlBodyAllList)
            else:
                # Excel空的时候，值为None
                sqlBodyAllList.append('')
                # nullCheckList.append()
                # print("读取到的SQL文各项目值错误")
        # 一行的数据不全为空的数据
        if (nullCheck) :
            sqlBodyList.append(sqlBodyAllList)
    return sqlBodyList
    print('')
    
def readExcelmain(paraUpDirectory):
    if (paraUpDirectory == '') :
        #得到当前脚本的执行目录
        currentDirectory  =  os.getcwd()
        #查看是否已经存在备份目录，如果有则删除，没有则新建目录
        paraUpDirectory  =  "%s\\%s"  %( currentDirectory, "rule.xlsx")
    excel=load_workbook(paraUpDirectory, data_only=True)
    
    sheetNameList = getSheetNames(excel)
    for sheetName in sheetNameList:
        if (sheetName == 'InputExcel') :
            print(sheetName)
            sqlList = getInputExcel(excel, str(sheetName))

def main(paraUpDirectory,excel):
    if (paraUpDirectory == '') :
        #得到当前脚本的执行目录
        currentDirectory  =  os.getcwd()
        #查看是否已经存在备份目录，如果有则删除，没有则新建目录
        paraUpDirectory  =  "%s\\%s"  %( currentDirectory, "SPUR Phase3_メンテンマスタ初期値設定リスト.xlsx")
    excel=load_workbook(paraUpDirectory, data_only=True)
    
    sheetNameList = getSheetNames(excel)
    for sheetName in sheetNameList:
        if (is_kazu(sheetName)) :
            print(sheetName)
            sqlList = getSheetDataList(excel, str(sheetName))


def getExcelSheetDatalist(excel,sheetName):
    ruleBodyList = {}
    table = excel.get_sheet_by_name(sheetName)
    # nulCheckList =[False,False]
    for x in range(1,table.max_row+1): #从2开始，第一行为title
        # 带遍烟历Excel表格列的内容
        rowData =(table.cell(row=x,column=1)).value #获取数据
        rulecolBodyAllList= []
        nullcheck = False
        for i in range(2,table.max_column+1): #从3开始，第一行为title
            # 看历Excel表格列的内容
            colData = (table.cell(row=x,column=i)).value #获取数据
            if(not(colData is None)):
                nullcheck = True
                rulecolBodyAllList.append(colData)
        # 一行的数据不全为空的数据
        if(nullcheck):
            ruleBodyList[rowData]=rulecolBodyAllList 
    return ruleBodyList

def setExcelData(ruleList, outList):
    # 取得Input文件名
    inputFilePath = ruleList['filename']
    # # 取得Output文件名
    outputFilePath = outList['filename']
    # 取得数据规则Head
    datarule = ruleList['datarule']
    inputDirectory = "%s\\%s" %( workDirectory,inputFilePath[0])
    inputExcel = load_workbook(inputDirectory, data_only=True)
    
    # Input信息取得
    workSheetList = []
    itemNameList = []
    itemList = []
    outputItemList = {}
    for rowstr in ruleList:
        # 取得Input Sheet名
        inputData = ruleList[rowstr]
        if 'workSheet' in rowstr:
            if (len(workSheetList) > 0):
                table = inputExcel.get_sheet_by_name(workSheetList[-1])
                outputItemList = getOutputValue(table,datarule,itemNameList,itemList,outputItemList)
                itemNameList = []
                itemList = []
            workSheetList.append(inputData[0])
        if 'item' in rowstr:
            # 字典是顺序操作的所以多sheet操作的时候按顺序操作即可，在处理到信sheet时需要创建新List
            itemNameList.append(rowstr)
            itemList.append(inputData)

    if (len(workSheetList) > 0):
        table = inputExcel.get_sheet_by_name(workSheetList[-1])
        outputItemList = getOutputValue(table,datarule,itemNameList,itemList,outputItemList)
        itemNameList = []
        itemList = []
    print(outputItemList)
    
    
    outputDirectory = "%s\\%s" %( workDirectory,outputFilePath[0])
    outputExcel = load_workbook(outputDirectory, data_only=True)
    # Output信息取得
    outWorkSheetList = []
    outItemNameList = []
    outItemList = []
    for outstr in outList:
        # 取得Input Sheet名
        outData = outList[outstr]
        if 'workSheet' in outstr:
            if (len(outWorkSheetList) > 0):
                table = outputExcel.get_sheet_by_name(outWorkSheetList[-1])
                setItemData(table,datarule,outItemNameList,outItemList,outputItemList)
                outItemNameList = []
                outItemList = []
            outWorkSheetList.append(outData[0])
        if 'item' in outstr:
            outItemNameList.append(outstr)
            outItemList.append(outData)
    if (len(outWorkSheetList) > 0):
        table = outputExcel.get_sheet_by_name(outWorkSheetList[-1])
        setItemData(table,datarule,outItemNameList,outItemList,outputItemList)
        outItemNameList = []
        outItemList = []
        updataoutputDirectory = "%s\\%s" %( workDirectory,'updata'+outputFilePath[0])
        # 信息写入后，保存成新文件
        outputExcel.save(updataoutputDirectory)
    # 根据项目ID取得对应项目值 设定到对应Output项目值
    
    
def Merge(dict1, dict2): 
    res = {**dict1, **dict2} 
    return res 

def getOutputValue(table,datarule,itemNameList,itemList,outputItemList):
    # *******************************
    # 获取各项目处理间的信息
    itmeValueList = getItemInfoValue(datarule,itemNameList,itemList)
    # 新的Sheet时，需要初始化list
    print('根据项目ID取得对应项目值')
    print(itmeValueList)
    # 对逐个项目的信息进行提取
    outputItemSheetList = getItemValue(table,itmeValueList)
    if (len(outputItemList) > 0):
        outputItemList = Merge(outputItemList, outputItemSheetList)
    else:
        outputItemList = outputItemSheetList
    # *******************************
    return outputItemList

# for i,rulestr in enumerate(datarule):

def setItemData(table,datarule,outItemNameList,outItemList,outputItemList):
    # 根据每个项目的取得信息，取得项目设定值
    
    # 获取各项目处理间的信息
    itmeValueList = getItemInfoValue(datarule,outItemNameList,outItemList)
    # 新的Sheet时，需要初始化list
    print('根据项目ID取得对应项目值')
    print(itmeValueList)
    for inkey,invalue in outputItemList.items():
    # for outItem in outputItemList:
        outitempath = itmeValueList[inkey]
        rownum = outitempath['rownum']
        colnum = outitempath['colnum']
        (table.cell(row=int(rownum),column=int(colnum))).value = invalue


# 对逐个项目的信息进行提取
def getItemValue(table,itmeValueList):
    outputItemList = {}
    # 项目根据项目关联信息逐个取得结果值
    for inkey,invalue in itmeValueList.items():
        outputItemList[inkey] = getOneInfoValue(invalue,table)
        # print(outputItemList[inkey])
    print(outputItemList)
    return outputItemList
    
# 单一单元格设定值取得
def getOneInfoValue(invalue,table):
    # {'rownum': 24, 'colnum': 3, 'iftype': 'merge', 'setvalue': ',', 'condition': 'none'}
    outValue = ''
    rownumList = []
    colnumList = []
    iftypeList = []
    setvalueList = []
    conditionList = []
    for inkey,invalue in invalue.items():
        if ('rownum' in inkey):
            rownumList.append(invalue)
        if ('colnum' in inkey):
            colnumList.append(invalue)
        if ('iftype' in inkey):
            iftypeList.append(invalue)
        if ('setvalue' in inkey):
            setvalueList.append(invalue)
        if ('condition' in inkey):
            conditionList.append(invalue)
    if (len(rownumList) > 1):
        outValueTmp = ''
        for i in range(len(rownumList)):
            # 取得项目值
            rownumvalue = rownumList[i]
            colnumvalue = colnumList[i]
            rowData = (table.cell(row=int(rownumvalue),column=int(colnumvalue))).value #获取数据
            # 取得条件类型 
            iftypevalue = checkListNull(iftypeList, i)
            # 取得条件判断后的设定值
            setvaluevalue = checkListNull(setvalueList, i)
            # 取得条件判断设定值
            conditionvalue = checkListNull(conditionList, i)
            outValueTmp = getValue(rowData,iftypevalue,setvaluevalue,conditionvalue)
            outValue += outValueTmp
    else:
        rownumvalue = rownumList[0]
        colnumvalue = colnumList[0]
        rowData = (table.cell(row=int(rownumvalue),column=int(colnumvalue))).value #获取数据
        # 取得条件类型 
        iftypevalue = checkListNull(iftypeList, 0)
        # 取得条件判断后的设定值
        setvaluevalue = checkListNull(setvalueList, 0)
        # 取得条件判断设定值
        conditionvalue = checkListNull(conditionList, 0)
        outValue = getValue(rowData,iftypevalue,setvaluevalue,conditionvalue)
    print(outValue)
    return outValue

def checkListNull(paraList,index):
    putvalue = ''
    if (len(paraList) > index):
        putvalue = paraList[index]
    else :
        putvalue = ''
    return putvalue
# 根据条件设定出力值
# iftype ：merge	单纯合并文字列，合并目标setvalue,合并目标rownum1和colnum1的值，如果有rownum2和colnum2也要合并
#          mergestr	单纯合并文字列，合并目标setvalue
#          greater	大于condition的值时，设定setvalue的值
# setvalue：赋值文字列
# condition：条件值 如果是none视为无效
def getValue(rowData,iftypevalue,setvaluevalue,conditionvalue):
    if (iftypevalue == 'merge'):
        return str(rowData) + str(setvaluevalue)
    elif (iftypevalue == 'mergestr'):
        return str(rowData) + str(setvaluevalue)
    elif (iftypevalue == 'greater'):
        if (float(rowData) > float(conditionvalue)):
            return setvaluevalue
    elif (iftypevalue == 'less'):
        if (float(rowData) < float(conditionvalue)):
            return setvaluevalue
    else:
        return rowData
    print('')
    
# 获取各项目处理间的信息，供后续处理调用
def getItemInfoValue(datarule,itemNameList,itemList):
    itmeValue = {}
    for itemnmae,item in zip(itemNameList,itemList):
        itmeValuetmp={}
        itmeValuelist=[]
        itmeValuetmp[itemnmae] = zip_longest(datarule,item)
        itemonelist = itmeValuetmp[itemnmae]
        for itemone in itemonelist:
            if (itemone[1] != None):
                itmeValuelist.append(itemone)
                # del itemonelist[itemone]
        itmeValue[itemnmae] = dict(itmeValuelist)
    print(itmeValue)
    return itmeValue

# 'datarule': ['rownum', 'colnum', 'iftype', 'setvalue', 'condition', 'rownum', 'colnum', 'iftype', 'rownum', 'colnum']
# 'item3': [20, 3], 
# 'item4': [21, 3, 'and', ',', 'none', 23, 3, 'and', 24, 3], 
# 'item5': [22, 3, ' County'], 
# 'item6': [29, 3],

# 指定内容取得
def getData(table,rowNum,colNum) :
    strData = (table.cell(row=rowNum,column=colNum)).value #获取数据
    return strData

def maintest(paralpDirectory,excel):
    excel=load_workbook(paralpDirectory, data_only=True)

    sheetNameList = getSheetNames(excel)
    ruleList=[]
    outList = []
    for sheetName in sheetNameList:
        if("excel"in sheetName):
            print(sheetName)
            if('input'in sheetName):
                ruleList = getExcelSheetDatalist(excel,str(sheetName))
                print('ruleList')
                print(ruleList)
            if('output' in sheetName):
                outList = getExcelSheetDatalist(excel, str(sheetName))
                print('outList')
                print(outList)
            # Input文件信息取得
    setExcelData(ruleList,outList)
    
workDirectory = ''
# if  __name__  ==  '__main__' :
#     invalue = {'rownum': 21, 'colnum': 3, 'iftype': 'merge', 'setvalue': ',', 'condition': 'none', 'rownum1': 23, 'colnum1': 3, 'iftype1': 'merge', 'setvalue1': '', 'condition1': 'none', 'rownum2': 24, 'colnum2': 3, 'iftype2': 'none', 'setvalue2': 'none', 'condition2': 'none'}
#     getOneInfoValue(invalue)
if  __name__  ==  '__main__' :
    
    # 得到当前脚本的执行目录
    cunnentDirectory = os.getcwd()
    tempPath = "doc\\input"
    workDirectory = "%s\\%s" %( cunnentDirectory,tempPath)
    # 希查看是否已经存在备份目录，如果有则删除，没有则新建目录
    paralpDirectory = "%s\\%s\\%s" %( cunnentDirectory,tempPath,"rule.xlsx")
    maintest(paralpDirectory,excel)
    # getSheetDataList("1")

#     print ("\n读取到完成...\n回车键退出")

# if  __name__  ==  '__main__' :
#     datarule = ['rownum', 'colnum', 'iftype', 'setvalue', 'condition', 'rownum', 'colnum', 'iftype', 'rownum', 'colnum']
#     itemNameList = []
#     itemNameList.append('item3')
#     itemNameList.append('item4')
#     itemNameList.append('item5')
#     itemNameList.append('item6')
#     itemList = []
#     itemList.append([20, 3])
#     itemList.append([21, 3, 'and', ',', 'none', 23, 3, 'and', 24, 3])
#     itemList.append([22, 3, ' County']) 
#     itemList.append([29, 3])
#     getItemValue(datarule,itemNameList,itemList)
